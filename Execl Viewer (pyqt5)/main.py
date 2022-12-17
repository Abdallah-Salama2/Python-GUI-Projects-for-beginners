from PyQt5.QtWidgets import *
import sys
import openpyxl


class Main(QWidget):
    def __init__(self):
        super(Main,self).__init__()
        self.setWindowTitle("Load Excel data to QtableWidget")

        layout = QVBoxLayout()
        self.setLayout(layout)

        self.table_Widget = QTableWidget()
        layout.addWidget(self.table_Widget)

        self.load_data()


    def load_data(self):
        path = "list-countries-world.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active #refer to active sheet in my excel cuz excel file have one or tow sheets he pick the active one

        self.table_Widget.setRowCount(sheet.max_row)
        self.table_Widget.setColumnCount(sheet.max_column)

        values = list(sheet.values) # values get me all differnet values in sheet (list of tuples with my values)
        for value in values:
            print(value)  # sample of o/p :
        ''''# ('Rank', 'Country ', 'Population', '% of world pop.')
            (0, 'World', 7135000000, 1)
            (1, 'China', 1362010000, 0.191)
            (2, 'India', 1238570000, 0.174)
            (3, 'United States', 317400000, 0.0445)'''

        self.table_Widget.setHorizontalHeaderLabels(values[0]) # load data into Qtable widget but notice to set Rows and column count
        ## we added the header 4 columns with titles and all rows in sheet  but all are empty
        ## we need to populate the tabel

        #self.table_Widget.setItem(0,2,QTableWidgetItem("Hello")) added hello at first row(1) column 3 (population here)
        row_index=0
        for value_tuple in values[1:]: #started from 1 so he dont take header columns with labels
            col_index=0
            for value in value_tuple:
                self.table_Widget.setItem(row_index,col_index,QTableWidgetItem(str(value)))
                col_index +=1

            row_index +=1
        # hy5osh 3ala kol row y7ot data fe kol col 5als awl col y2om yzwd l7ad ma y5ls cols bta3t row kolo ba3den
        # yro7 3ala tany row y5ls cols feh w hakaza


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Main() ## Window variable with instance of main class
    window.showMaximized()
    app.exec_()